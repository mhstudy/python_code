# -*- coding: utf-8 -*-
import os  # 用于操作文件和目录
import csv  # 用于读写 CSV 格式的文件
import enum  # 用于定义枚举类型
import json  # 用于处理 JSON 数据
import chardet  # 用于检测文本文件的编码格式。需要安装 pip install chardet。
import requests  # 用于发送 HTTP 请求，并处理响应。需要安装 pip install requests。
import openpyxl  # 用于读写 Excel 文件。需要安装 pip install openpyxl。
import pandas as pd  # 用于处理数据。需要安装 pip install pandas。
import mysql.connector  # 用于连接 MySQL 数据库和执行 SQL 语句。需要安装 pip install mysql-connector-python。
from bs4 import BeautifulSoup  # 用于解析 HTML 和 XML 文件，并提取其中的数据。需要安装 pip install beautifulsoup4。
from dateutil.parser import parse  # 用于解析日期和时间字符串。需要安装 pip install python-dateutil。
from datetime import datetime, timedelta  # 用于处理日期和时间。


class OddsCompany(enum.Enum):
    """
    枚举类，定义了三个选项，每个选项都有一个cid和prefix属性，cid是一个整数，prefix是一个字符串。
    """
    COMPANY_OFFICIALLOTTERY = (1, 'jingcai')
    COMPANY_BET365 = (3, 'bet365')
    COMPANY_AOMEN = (5, 'MB')

    def __init__(self, cid, prefix):
        """
        构造函数，初始化选项的cid和prefix属性
        """
        self.cid = cid
        self.prefix = prefix

    @classmethod
    def get_prefix_by_cid(cls, cid):
        """
        类方法，接受一个cid作为参数，并返回与该cid对应的选项。
        如果没有cid对应的选项，则抛出ValueError
        """
        for company in cls:
            if company.cid == cid:
                return company.prefix
        raise ValueError(f"No company with cid={cid}")


# 足球-odd
class FootBallOdd:
    def __init__(self, timeout=10, proxies=None):
        """
        初始化 FootBallOdd 对象。

        Args:
            timeout (int, optional): 请求超时时间，单位为秒。默认为 10。
            proxies (dict, optional): 代理服务器设置。默认为 None。
        """
        # 设置默认请求头
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        # 创建 requests.Session 对象
        self.session = requests.Session()
        # 更新默认请求头和 session 的请求头
        self.session.headers.update(self.headers)
        # 设置超时时间
        self.session.timeout = timeout
        # 设置代理服务器
        self.session.proxies = proxies

    # 获取赔率信息
    def get_odds(self, fid, cid=OddsCompany.COMPANY_BET365.cid):
        """
        获取给定赛事和公司的赔率值。

        Args:
            fid : 赛事 ID。
            cid : 公司 ID，默认为 BET365  此处只有 BET365 和 澳门 两个值有用

        Returns:
            odds_dict: 给定赛事的公司的赔率值。

        """
        # https://live.500.com/json/odds.php?fid=1077272&cid=3
        url = f'https://live.500.com/json/odds.php?fid={fid}&cid={cid}'

        response = self.session.get(url)
        response.encoding = response.apparent_encoding

        # 将字符串解析成 Python 中的列表和字典等数据结构
        data = json.loads(response.text.strip())

        # 添加前缀
        prefix = OddsCompany.get_prefix_by_cid(cid)

        try:
            odds_dict = {
                prefix + '_yazhi_end': data[0][1],
                prefix + '_yazhi_start': data[0][0],
                prefix + '_ouzhi_end': data[1][1],
                prefix + '_ouzhi_start': data[1][0],
                prefix + '_daxiao_end': data[2][1],
                prefix + '_daxiao_start': data[2][0]
            }
        except (TypeError, IndexError):
            odds_dict = {
                prefix + '_yazhi_end': '',
                prefix + '_yazhi_start': '',
                prefix + '_ouzhi_end': '',
                prefix + '_ouzhi_start': '',
                prefix + '_daxiao_end': '',
                prefix + '_daxiao_start': ''
            }
        return odds_dict

    # 完场比分-比赛fid
    def get_wangchang(self, date_time=None):
        """
           获取指定日期或当前日期的所有完场比赛的信息。

           Args:
               date_time (datetime, optional): 要查询的日期时间，如果未指定，则使用当前日期时间。默认为 None。

           Returns:
               list: 所有完场比赛的比分信息，包含比赛 ID、主队名称、客队名称和比分等
        """
        if date_time is None:
            date_time = datetime.today() - timedelta(days=1)
        else:
            # date_time = datetime.strptime(str(date_time), '%Y%m%d').strftime('%Y-%m-%d')
            date_time = parse(str(date_time)).strftime('%Y-%m-%d')

        # https://live.500.com/wanchang.php?e=2023-06-29
        url = f'https://live.500.com/wanchang.php?e={date_time}'

        # 打印数据来源网址
        print(url)

        response = self.session.get(url)
        response.encoding = response.apparent_encoding

        soup = BeautifulSoup(response.text, 'html.parser')

        div = soup.find('div', {'class': 'wrap'}) \
            .find('table', {'id': 'table_match'}) \
            .find('tbody')

        my_list = []

        for span in div.find_all('span', {'class': 'red'}):
            # 检查span标签的文本是否为"完"
            if span.text == '完':
                # 输出包含"完"的那行数据所在的<tr>标签
                trs = span.parent.parent

                dict = {
                    'ds': date_time,
                    'fid': trs.get('id')[1:],
                    'competition': trs.get('gy').split(',')[0],
                    'rounds': trs.find_all('td', {'align': 'center'})[0].text.strip(),
                    'competition_time': trs.find_all('td', {'align': 'center'})[1].text.strip(),
                    'mainName': trs.get('gy').split(',')[1],
                    'clientName': trs.get('gy').split(',')[2],
                    'half_time_score': '{}:{}'.format(*trs.find('td', {'class': 'red'}).text.strip().split(' - ')),
                    'full_time_score': trs.find('div', {'class': 'pk'}).select('.pk a')[0].text.strip() + ':'
                                       + trs.find('div', {'class': 'pk'}).select('.pk a')[2].text.strip(),
                }
                my_list.append(dict)
        return my_list

    # 获取完整赔率信息
    def get_wangchang_odds(self, date_time=None):
        """
        获取完整赔率信息。

        Args:
            date_time (str, optional): 过滤赔率的日期时间，格式为 'YYYY-MM-DD'。不传默认为当前时间-1天

        Returns:
            dict: 包含完整赔率信息的字典。字典的键为赔率类型，值为包含赔率信息的列表。赔率类型包括 'yazhi_end'（亚洲终盘赔率）、'yazhi_start'（亚洲初盘赔率）、'ouzhi_end'（大小球终盘赔率）、'ouzhi_start'（大小球初盘赔率）、'daxiao_end'（大小球终盘大小）、'daxiao_start'（大小球初盘大小）。

        Raises:
            ValueError: 如果提供的日期时间格式不正确，将引发 ValueError 异常。
            ConnectionError: 如果无法连接到赔率数据源，将引发 ConnectionError 异常。
        """
        if date_time is None:
            date_time = datetime.today() - timedelta(days=1)
        else:
            # date_time = datetime.strptime(str(date_time), '%Y%m%d').strftime('%Y-%m-%d')
            date_time = parse(str(date_time)).strftime('%Y-%m-%d')

        result_list = []

        wangchang_list = self.get_wangchang(date_time)

        for wangchang_dict in wangchang_list:
            odds_dict = self.get_odds(wangchang_dict['fid'], cid=OddsCompany.COMPANY_BET365.cid)
            jingcai_dict = self.ouzhi(wangchang_dict['fid'], cid=OddsCompany.COMPANY_OFFICIALLOTTERY.cid)
            result_dict = {**wangchang_dict, **odds_dict, **jingcai_dict}
            result_list.append(result_dict)

        return result_list

    def ouzhi(self, fid, cid):
        # 'https://odds.500.com/fenxi/ouzhi-1078507.shtml'
        url = f'https://odds.500.com/fenxi/ouzhi-{fid}.shtml'

        response = self.session.get(url)
        response.encoding = response.apparent_encoding

        soup = BeautifulSoup(response.text, 'html.parser')

        # 获取比赛名称和比赛时间
        div = soup.find('div', {'class': 'odds_hd_cont'})
        mainName = div.select('a.hd_name')[0].text
        clientName = div.select('a.hd_name')[2].text
        league = div.select('a.hd_name')[1].text.strip()
        score = '' if div.select_one('p.odds_hd_bf').get_text(strip=True) == 'VS' \
            else div.select_one('p.odds_hd_bf').get_text(strip=True)
        game_time = str(div.select_one('p.game_time').text)[4:]

        try:
            # 获取cid的赔率信息
            jingcai = soup.find('div', {'class': 'odds_content'}) \
                .find('table', {'id': 'datatb'}) \
                .find('tr', {'id': str(cid)})
            jingcai_tds = jingcai.select('table.pl_table_data')[0].find_all('td')
            mylist = []
            for td1 in jingcai_tds:
                mylist.append(td1.text.strip())
        except AttributeError as e:
            # 处理查找标签时可能出现的异常
            mylist = ['', '', '', '', '', '', '']

        # 前缀
        prefix = OddsCompany.get_prefix_by_cid(cid)

        # 将赔率信息存储在一个字典中
        odds_dict = {
            'mainName': mainName,
            'clientName': clientName,
            'full_time_score': score,
            'game_time': game_time,
            'league': league,
            prefix + '_ouzhi_end': [mylist[3], mylist[4], mylist[5]] \
                if mylist[3] != '' or mylist[4] != '' or mylist[5] != '' else '',
            prefix + '_ouzhi_start': [mylist[0], mylist[1], mylist[2]] \
                if mylist[0] != '' or mylist[1] != '' or mylist[2] != '' else ''
        }
        return odds_dict


class MySQLPool:
    """
    一个表示连接到MySQL数据库的连接池的类。

    参数:
        user (str): MySQL的用户名。
        password (str): 给定用户的MySQL密码。
        database (str): 要连接的数据库的名称。
        host (Optional[str]): 数据库服务器的主机名。默认为'localhost'。
        port (Optional[int]): 连接使用的端口号。默认为3306。
        pool_size (Optional[int]): 连接池中允许的最大连接数。默认为5。
        charset (Optional[str]): 连接使用的字符集。默认为'utf8mb4'。
        autocommit (Optional[bool]): 是否自动提交更改到数据库。默认为False。
        connect_timeout (Optional[int]): 等待连接的秒数。默认为30。

    属性:
        user (str): MySQL的用户名。
        password (str): 给定用户的MySQL密码。
        database (str): 要连接的数据库的名称。
        host (str): 数据库服务器的主机名。
        port (int): 连接使用的端口号。
        pool_size (int): 连接池中允许的最大连接数。
        charset (str): 连接使用的字符集。
        autocommit (bool): 是否自动提交更改到数据库。
        connect_timeout (int): 等待连接的秒数。
        pool_args (dict): 包含传递给MySQLConnectionPool构造函数的参数的字典。
        pool (mysql.connector.pooling.MySQLConnectionPool): 连接池对象。

    方法:
        get_connection(): 从连接池中返回一个连接。
        close_connection(conn, cursor=None): 关闭连接，可选地关闭游标。
    """

    def __init__(self, user, password, database, host='localhost', port=3306, pool_size=5, charset='utf8mb4',
                 autocommit=False,
                 connect_timeout=30):
        self.user = user
        self.password = password
        self.database = database
        self.host = host
        self.port = port
        self.pool_size = pool_size
        self.charset = charset
        self.autocommit = autocommit
        self.connect_timeout = connect_timeout

        self.pool_args = {
            'pool_size': self.pool_size,
            'host': self.host,
            'port': self.port,
            'user': self.user,
            'password': self.password,
            'database': self.database,
            'charset': self.charset,
            'autocommit': self.autocommit,
            'connect_timeout': self.connect_timeout
        }

        self.pool = mysql.connector.pooling.MySQLConnectionPool(**self.pool_args)

    def get_connection(self):
        """
        从连接池中获取连接。

        返回:
            mysql.connector.connection.MySQLConnection: 从连接池中获取的连接对象。
        """
        return self.pool.get_connection()

    def close_connection(self, conn, cursor=None):
        """
        关闭连接和游标。

        参数:
            conn (mysql.connector.connection.MySQLConnection): 要关闭的连接对象。
            cursor (Optional[mysql.connector.cursor.MySQLCursor]): 要关闭的游标对象。默认为None。
        """
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()


class DataUtils:
    @staticmethod
    def get_date_range(start_time, end_time):
        """
        获取指定日期范围内的所有日期字符串列表。

        Args:
            start_time (int): 起始日期，格式为 YYYYMMDD。
            end_time (int): 结束日期，格式为 YYYYMMDD。

        Returns:
            List[str]: 所有日期的字符串列表，格式为 YYYY-MM-DD。

        Raises:
            ValueError: 如果起始日期或结束日期无效，则引发此异常。
        """
        start_date = datetime.strptime(str(start_time), '%Y%m%d')
        end_date = datetime.strptime(str(end_time), '%Y%m%d')
        date_diff = end_date - start_date
        date_list = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(date_diff.days + 1)]
        return date_list

    @staticmethod
    def read_csv(file_name):
        """
        读取 CSV 文件，并将其转换为字典列表。

        Args:
            file_name (str): CSV 文件路径。

        Returns:
            list: 包含 CSV 文件内容的字典列表。

        Raises:
            UnicodeDecodeError: 如果文件编码无法解码文件中的字符。
        """
        # 使用 chardet 库检测文件编码
        with open(file_name, 'rb') as f:
            result = chardet.detect(f.read())
        # 使用检测结果打开 CSV 文件
        with open(file_name, mode='r', encoding=result['encoding']) as csv_file:
            reader = csv.DictReader(csv_file)
            return [row for row in reader]

    @staticmethod
    def write_to_csv(dict_list, file_name='test.csv', mode='w'):
        """
        将字典列表写入 CSV 文件。

        Args:
            dict_list (List[Dict[str, Any]]): 包含要写入的数据的字典列表。
            file_name (str): CSV 文件路径。默认为 'test.csv'。
            mode (str): 写入模式。默认为写入模式（'w'），也可以指定为追加模式（'a'）。
        """
        # 获取字典列表中第一个字典的键名列表
        keys = dict_list[0].keys()

        # 打开指定的 CSV 文件，指定写入模式和换行符参数
        with open(file_name, mode=mode, encoding='utf-8', newline='') as csv_file:
            # 创建一个 DictWriter 对象，指定字段名为 keys
            writer = csv.DictWriter(csv_file, fieldnames=keys)
            if mode == 'w':
                # 写入 CSV 文件的表头（即字段名）
                writer.writeheader()
            # 遍历字典列表中的每个字典，将其写入 CSV 文件
            for row in dict_list:
                writer.writerow(row)

    @staticmethod
    def read_json(file_name):
        """
        从 JSON 文件中读取数据

        参数：
            file_name -- 文件名

        返回：
            JSON 文件中的数据

        """
        # 使用 chardet 模块自动识别文件编码
        with open(file_name, 'rb') as f:
            result = chardet.detect(f.read())
            encoding = result['encoding']

        # 使用 open 函数打开指定的文件，模式为 'r'，使用识别出的编码
        with open(file_name, mode='r', encoding=encoding) as json_file:
            # 使用 json.load 函数从文件中读取数据
            return json.load(json_file)

    @staticmethod
    def write_to_json(data, file_name='test.json', mode='w'):
        """
        将数据写入 JSON 文件中

        Args：
            data -- 要写入的数据
            file_name -- 文件名，默认为 'test.json'
            mode -- 打开文件的模式，默认为 'w'，表示覆盖写入；设置为 'a' 则表示追加写入
        """
        # 使用 open 函数打开指定的文件，模式为 mode，编码格式为 UTF-8
        with open(file_name, mode=mode, encoding='utf-8') as json_file:
            # 使用 json.dump 函数将数据写入文件中，其中 ensure_ascii 参数设置为 False，可以保证中文字符正确写入文件
            json.dump(data, json_file, ensure_ascii=False)

    @staticmethod
    def write_to_excel(data, filename='text.xlsx', mode='w'):
        """
        将字典列表写入 Excel 文件。

        Args：
            data -- 要写入的数据
            file_name -- 文件名，默认为 'test.xlsx'
            mode -- 打开文件的模式，默认为 'w'，表示覆盖写入；设置为 'a' 则表示追加写入
        """
        # 校验文件打开模式
        if mode not in ['w', 'a']:
            raise ValueError("文件打开模式必须是 'w' 或 'a'")

        # 判断文件是否存在
        file_exists = os.path.isfile(filename)

        # 如果文件已经存在并且打开模式为 'a'，则打开现有的工作簿
        if file_exists and mode == 'a':
            try:
                # 打开现有的工作簿
                wb = openpyxl.load_workbook(filename)
                # 选择默认的工作表
                ws = wb.active
                # 找到最后一行的行号
                last_row = ws.max_row
                # 从最后一行的下一行开始写入数据
                start_row = last_row + 1
            except PermissionError:
                print(f"文件 '{filename}' 正在被其他程序使用，请稍后重试...")
                return
        else:
            # 否则，创建一个新的工作簿
            wb = openpyxl.Workbook()
            # 选择默认的工作表
            ws = wb.active
            # 写入第一行，即字典键
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            # 从第二行开始写入数据
            start_row = 2

        # 写入数据
        for row_num, row_data in enumerate(data, start_row):
            for col_num, cell_value in enumerate(row_data.values(), 1):
                # 将空字符串转换为 None
                if cell_value == '':
                    cell_value = None
                # 将列表转换为字符串，并使用逗号分隔项
                elif isinstance(cell_value, list):
                    cell_value = ', '.join(str(x) for x in cell_value)
                ws.cell(row=row_num, column=col_num, value=cell_value)

        # 保存工作簿
        wb.save(filename)

    @staticmethod
    def read_excel(file_name, sheet_name=None):
        """
        读取 Excel 文件并返回 pandas DataFrame。

        Args:
            file_name (str): 要读取的 Excel 文件的文件名。
            sheet_name (str, optional): 要从 Excel 文件中读取的工作表的名称。
                如果为 None，则读取文件中的第一个工作表。

        Returns:
            pd.DataFrame: 包含指定工作表数据的 pandas DataFrame。

        Raises:
            ValueError: 如果在 Excel 文件中找不到指定的工作表名称。
        """
        sheets = pd.read_excel(file_name, sheet_name=None)
        sheet_name = next(iter(sheets)) if sheet_name is None else sheet_name
        if sheet_name not in sheets:
            raise ValueError(f"在文件 '{file_name}' 中找不到工作表 '{sheet_name}'")
        return pd.read_excel(file_name, sheet_name)

    def __init__(self, user, password, database, host='localhost', port=3306):
        # 初始化连接池
        self.pool = MySQLPool(user=user, password=password, database=database, host=host, port=port)

    def execute_query(self, query, params=None):
        # 执行一个查询并返回结果
        conn = None
        cursor = None
        try:
            # 获取连接和游标对象
            conn = self.pool.get_connection()
            cursor = conn.cursor()
            # 执行查询
            cursor.execute(query, params)
            # 获取查询结果
            result = cursor.fetchall()
            return result
        finally:
            # 关闭连接和游标
            MySQLPool.close_connection(conn, cursor)

    def execute_update(self, query, params=None):
        # 执行一个更新并返回受影响的行数

        conn = None
        cursor = None
        try:
            # 获取连接和游标对象
            conn = self.pool.get_connection()
            cursor = conn.cursor()
            # 执行更新
            cursor.execute(query, params)
            # 获取受影响的行数
            affected_rows = cursor.rowcount
            # 提交更改
            conn.commit()
            return affected_rows
        finally:
            # 关闭连接和游标
            MySQLPool.close_connection(conn, cursor)

    def insert_data(self, table_name, data_list):
        """
        批量插入数据到MySQL表

        Args:
            table_name: MySQL表名
            data_list: 包含多个字典的列表，每个字典表示一行数据
        """
        conn = None
        cursor = None
        try:
            # 获取数据库连接和游标对象
            conn = self.pool.get_connection()
            cursor = conn.cursor()

            # 构造插入语句和更新表达式
            keys = data_list[0].keys()  # 获取字典的键
            columns = ', '.join(keys)  # 将键拼接成字符串
            placeholders = ', '.join(['%s'] * len(keys))  # 构造占位符字符串
            query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE"
            update_expr = ', '.join([f"{key}=VALUES({key})" for key in keys])  # 构造更新表达式
            query = f"{query} {update_expr}"  # 将更新表达式拼接到插入语句后面

            # 构造数据列表并批量插入数据
            values = []
            for data in data_list:
                row_values = []
                for value in data.values():
                    if isinstance(value, list):
                        # 如果值是列表，将其转换为JSON字符串
                        row_values.append(json.dumps(value, ensure_ascii=False))
                    else:
                        row_values.append(value)
                values.append(tuple(row_values))
            cursor.executemany(query, values)  # 批量插入数据

            # 提交事务
            conn.commit()
        except mysql.connector.IntegrityError as e:
            print(f"Error: {e}")
            conn.rollback()  # 回滚事务
        finally:
            MySQLPool.close_connection(conn, cursor)  # 关闭数据库连接和游标对象


# 获取一天的数据
def write_football_data_to_mysql_1d(date_time=None):
    # 创建FootBallOdd实例
    football_odd = FootBallOdd()

    # 创建DataUtils实例，并连接到MySQL
    data_utils = DataUtils(user='root', password='000000', database='football')

    # 结果字典
    result_list = football_odd.get_wangchang_odds(date_time)

    # 将数据插入到MySQL表中
    data_utils.insert_data(table_name='football_odd_data', data_list=result_list)


# 获取n天的数据
def write_football_data_to_mysql_nd(start_time, end_time):
    # # 写法一: 此写法从性能角度来看是最好的。
    # # 创建FootBallOdd实例
    # football_odd = FootBallOdd()
    #
    # # 创建DataUtils实例，并连接到MySQL
    # data_utils = DataUtils(user='root', password='000000', database='football')
    #
    # # 生成起始时间和结束时间之间的日期列表
    # date_list = DataUtils.get_date_range(start_time, end_time)
    # for date_time in date_list:
    #     result_list = football_odd.get_wangchang_odds(date_time)
    #     # 将数据插入到MySQL表中
    #     data_utils.insert_data(table_name='football_odd_data', data_list=result_list)

    # 写法二:
    # 生成起始时间和结束时间之间的日期列表
    date_list = DataUtils.get_date_range(start_time, end_time)
    for date_time in date_list:
        write_football_data_to_mysql_1d(date_time)


# 写入数据到csv文件
def write_football_date_to_csv_1d(football_odd: FootBallOdd, date_time=None, file_name='test.csv', mode='w'):
    result_list = football_odd.get_wangchang_odds(date_time)
    DataUtils.write_to_csv(result_list, file_name=file_name, mode=mode)


if __name__ == '__main__':
    write_football_data_to_mysql_nd(20230916, 20240108)

    # 写入csv文件
    # football_odd = FootBallOdd()

    # write_football_date_to_csv_1d(football_odd, date_time=20230707, file_name='test.csv', mode='w')

    # date_list = DataUtils.get_date_range(20230701, 20230706)
    #
    # # print(date_list)
    # for date_time in date_list:
    #     # print(date_time)
    #     write_football_date_to_csv_1d(football_odd, date_time, file_name='test.csv', mode='a')

    pass
